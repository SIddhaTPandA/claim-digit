"""
Excel Generator — Phase 1 SPD Benefits Extraction.

Generates the standardized 15-column Excel output from BenefitRecord objects.
Applies professional formatting: header fill, category highlighting, borders,
auto-column widths, and a frozen header row.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.models.benefit_record import BenefitRecord

logger = logging.getLogger(__name__)



_JSON_EXCEL_COLUMNS = [
    "Header", "Service",
    "In-Network Coinsurance", "In-Network After Deductible Flag", "In-Network Copay",
    "Out-Of-Network Coinsurance", "Out-Of-Network After Deductible Flag", "Out-Of-Network Copay",
    "Individual In-Network", "Family In-Network",
    "Individual Out-Of-Network", "Family Out-Of-Network",
    "Limit Type", "Limit Period",
    "Pre-Authorization Required", "Confidence Score",
]

class ExcelGenerator:
    """
    Writes a list of BenefitRecord objects to an Excel workbook.

    Output sheet "Benefits" has the 15 standard columns.
    An optional "Metadata" sheet is appended when *include_metadata=True*.
    """

    COLUMN_HEADERS: List[str] = BenefitRecord.get_excel_headers()

    # Styles
    _HEADER_FILL    = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
    _CATEGORY_FILL  = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    _CATEGORY_FONT  = Font(bold=True, size=10)
    _BORDER         = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def __init__(self, output_path: str) -> None:
        self.output_path = output_path
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # ----------------------------------------------------------------- public

    def generate(
        self,
        records: List[BenefitRecord],
        include_metadata: bool = True,
    ) -> str:
        """
        Write *records* to the Excel file and return the output path.

        Args:
            records: Normalized BenefitRecord objects.
            include_metadata: Whether to add a summary Metadata sheet.

        Returns:
            Absolute path to the generated Excel file.
        """
        rows = [r.to_excel_row() for r in records]
        df   = pd.DataFrame(rows, columns=self.COLUMN_HEADERS)

        with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Benefits")
            if include_metadata:
                self._write_metadata_sheet(writer, records)

        self._apply_formatting()
        logger.info("Excel written: %s (%d records)", self.output_path, len(records))
        return self.output_path

    # ------------------------------------------------------------ formatting

    def _apply_formatting(self) -> None:
        wb = load_workbook(self.output_path)
        ws = wb["Benefits"]

        # Header row
        for cell in ws[1]:
            cell.fill      = self._HEADER_FILL
            cell.font      = self._HEADER_FONT
            cell.border    = self._BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data rows
        current_header = None
        for row_idx in range(2, ws.max_row + 1):
            hdr_cell = ws.cell(row=row_idx, column=1)
            if hdr_cell.value and hdr_cell.value != current_header:
                current_header   = hdr_cell.value
                hdr_cell.fill    = self._CATEGORY_FILL
                hdr_cell.font    = self._CATEGORY_FONT

            for col_idx in range(1, ws.max_column + 1):
                cell           = ws.cell(row=row_idx, column=col_idx)
                cell.border    = self._BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Auto column widths (capped at 40)
        for col in ws.columns:
            max_len = 0
            letter  = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"
        wb.save(self.output_path)

    # --------------------------------------------------------- metadata sheet

    def _write_metadata_sheet(
        self,
        writer: pd.ExcelWriter,
        records: List[BenefitRecord],
    ) -> None:
        if not records:
            return

        avg_conf    = sum(r.confidence_score for r in records) / len(records)
        below_thr   = sum(1 for r in records if r.confidence_score < 0.70)
        unique_hdrs = len({r.header for r in records})

        meta = {
            "Property": [
                "Generation Date",
                "Total Records",
                "Unique Service Categories",
                "Average Confidence Score",
                "Records Below Threshold (0.70)",
            ],
            "Value": [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                len(records),
                unique_hdrs,
                f"{avg_conf:.2%}",
                below_thr,
            ],
        }
        pd.DataFrame(meta).to_excel(writer, index=False, sheet_name="Metadata")

    def generate_from_json(self, json_records, include_metadata=True):
        """
        Write flat-dict records produced by the CrewAI extraction crew.
        Each dict already contains the final column values keyed by the exact
        names in _JSON_EXCEL_COLUMNS. Missing fields are written as blank cells.
        """
        rows = []
        for rec in json_records:
            header  = str(rec.get("Header") or rec.get("header") or "General Services").strip()
            service = str(rec.get("Service") or rec.get("service") or "").strip()
            if not header or not service:
                continue
            row = {"Header": header, "Service": service}
            for col in _JSON_EXCEL_COLUMNS[2:]:
                row[col] = rec.get(col, "") or ""
            rows.append(row)

        if not rows:
            import logging as _l
            _l.getLogger(__name__).warning("generate_from_json: no valid rows to write.")
            return self.output_path

        import pandas as _pd
        df = _pd.DataFrame(rows, columns=_JSON_EXCEL_COLUMNS)
        with _pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Benefits")
            if include_metadata:
                self._write_json_metadata_sheet(writer, json_records)
        self._apply_formatting()
        logger.info("Excel written from JSON: %s (%d rows)", self.output_path, len(rows))
        return self.output_path

    def _write_json_metadata_sheet(self, writer, json_records):
        if not json_records:
            return
        unique_svcs = len({
            (str(r.get("Header") or r.get("header", "")),
             str(r.get("Service") or r.get("service", "")))
            for r in json_records
        })
        unique_hdrs = len({str(r.get("Header") or r.get("header", "")) for r in json_records})
        scores = []
        for r in json_records:
            try:
                scores.append(float(r.get("Confidence Score", 0) or 0))
            except (ValueError, TypeError):
                pass
        avg_conf = f"{sum(scores)/len(scores):.2%}" if scores else "N/A"
        from datetime import datetime as _dt
        import pandas as _pd
        meta = {
            "Property": [
                "Generation Date", "Total Records",
                "Unique Service Categories", "Unique Services",
                "Average Confidence Score", "Extraction Method",
            ],
            "Value": [
                _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                len(json_records), unique_hdrs, unique_svcs,
                avg_conf, "CrewAI / Azure OpenAI GPT",
            ],
        }
        _pd.DataFrame(meta).to_excel(writer, index=False, sheet_name="Metadata")
